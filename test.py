from openpyxl import load_workbook
# Specifying a sheet
wb = load_workbook(f"data/dwars_door_vlaanderen.xlsx")
ws = wb.active
max_col = ws.max_row
dwars = [0]*max_col
for i in range(max_col):
    dwars[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')


wb = load_workbook(f"data/flesh_wallon.xlsx")
ws = wb.active
max_col = ws.max_row
flesh = [0]*max_col
for i in range(max_col):
    flesh[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')





wb = load_workbook(f"data/milaan_sanremo.xlsx")
ws = wb.active
max_col = ws.max_row
milan = [0]*max_col
for i in range(max_col):
    milan[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')


wb = load_workbook(f"data/paris_roubaix.xlsx")
ws = wb.active
max_col = ws.max_row
paris = [0]*max_col
for i in range(max_col):
    paris[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')




wb = load_workbook(f"data/ronde_van_vlaanderen.xlsx")
ws = wb.active
max_col = ws.max_row
ronde = [0]*max_col
for i in range(max_col):
    ronde[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')


wb = load_workbook(f"data/omloop_het_nieuwsblad.xlsx")
ws = wb.active
max_col = ws.max_row
omloop = [0]*max_col
for i in range(max_col):
    omloop[i] = ws.cell(row=i+1, column=2).value.replace(u'\xa0', '')

# Print value of cell object
# using the value attribute
pento = []
for i in paris:
    if i in milan and i in flesh and i in dwars and i in ronde and i in omloop:
        pento.append(i)

print(pento)

# v2 = ws.range("F5").value
